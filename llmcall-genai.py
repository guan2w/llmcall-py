#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
LLM 批量调用：读取 Excel（prompt & QA），按 QA 的 Q 列逐行请求 LLM，
将 JSON 数组结果展开写回原文件。满足：
- 展开结果的所有行：Q 与 是否找到 相同
- 每个输入处理完成后立即落盘
- 支持 rows 范围、断点续跑、并发请求（请求并发，写入串行）

本版本使用 Google GenAI SDK (google-genai)
"""

import argparse
import datetime as dt
import json
import os
import re
import sys
import time
from typing import Any, Dict, List, Optional, Tuple

from google import genai
from google.genai import types
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# --- 列名配置 ---
COL_FOUND = "FOUND"  # 结果状态列名
COL_ERROR = "ERROR"      # 错误信息列名

# --- 配置解析（tomllib 优先） ---
try:
    import tomllib  # py311+
except ModuleNotFoundError:
    try:
        import tomli as tomllib  # py310-
    except Exception:
        tomllib = None


def log(msg: str) -> None:
    now = dt.datetime.now().strftime("%H:%M:%S")
    print(f"[{now}] {msg}", flush=True)


def mask_key_tail(key: Optional[str]) -> str:
    if not key:
        return "(empty)"
    tail = key[-5:] if len(key) >= 5 else key
    return "*" * max(0, len(key) - 5) + tail


def load_config(path: str) -> dict:
    if not os.path.exists(path):
        raise FileNotFoundError(f"配置文件不存在: {path}")
    if tomllib is None:
        raise RuntimeError("缺少 tomllib/tomli，请安装 tomli 或使用 Python 3.11+")
    with open(path, "rb") as f:
        return tomllib.load(f)


def merge_llm_config(cfg: dict, llm_name: str, cli_api_key: Optional[str]) -> dict:
    base = cfg.get("llm", {}) or {}
    # TOML 中 [llm.gemini_search] 会被解析为嵌套表 cfg["llm"]["gemini_search"]
    # 支持两种写法：
    # 1) [llm] + [llm.gemini_search] - 标准嵌套表
    # 2) cfg.get("llm.gemini_search", {}) - 容错（某些解析器可能支持）
    llm_section = cfg.get("llm", {}) or {}
    if isinstance(llm_section, dict):
        by_table = llm_section.get(llm_name, {})  # 标准方式：cfg["llm"]["gemini_search"]
    else:
        by_table = {}
    # 容错：尝试直接键访问（某些 TOML 解析器可能支持）
    if not by_table:
        by_table = cfg.get(f"llm.{llm_name}", {}) or {}

    # 合并：base <- by_table
    merged = dict(base)
    if isinstance(by_table, dict):
        merged.update(by_table)

    # CLI api_key 优先
    if cli_api_key:
        merged["api_key"] = cli_api_key

    # 必要字段检查
    # api_base 为可选，如果提供则用于自定义 API 端点
    api_base = merged.get("api_base")
    api_key = merged.get("api_key")
    model_id = merged.get("model_id")
    if not (api_key or merged.get("user_token")):
        raise ValueError("未提供 api_key（或 user_token）")
    if not model_id:
        raise ValueError("配置缺少 llm.model_id")

    # 默认并发/重试/超时
    merged.setdefault("parallel", 5)
    merged.setdefault("retry_times", 1)
    merged.setdefault("retry_delay", 10)
    merged.setdefault("timeout", 120)
    
    # 联网搜索功能（默认关闭）
    merged.setdefault("enable_google_search", False)
    
    # URL Context 功能（默认开启）
    merged.setdefault("enable_url_context", True)
    
    # 生成参数（可选）
    # temperature: 控制随机性，0.0-2.0，默认不设置（使用模型默认值）
    # thinking_budget: 思考预算，-1 表示无限制，默认不设置
    # 这些参数如果在配置中未设置，则不传递给 API（使用 API 默认值）
    
    return merged


def parse_rows_arg(rows_arg: Optional[str], data_start_row: int, data_end_row: int) -> List[int]:
    """
    rows 语法：
      - None: 处理 data_start_row..data_end_row
      - "2-5": 处理 2..5
      - "2+":  处理 2..data_end_row
    返回：原始行号列表（基于启动时的行号）
    """
    if not rows_arg:
        return list(range(data_start_row, data_end_row + 1))

    rows_arg = rows_arg.strip()
    m = re.fullmatch(r"(\d+)\-(\d+)", rows_arg)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        a = max(a, data_start_row)
        b = min(b, data_end_row)
        if a > b:
            return []
        return list(range(a, b + 1))

    m = re.fullmatch(r"(\d+)\+", rows_arg)
    if m:
        a = int(m.group(1))
        a = max(a, data_start_row)
        return list(range(a, data_end_row + 1))

    raise ValueError(f"rows 参数不合法: {rows_arg}")


def get_sheet(wb, name: str) -> Worksheet:
    if name not in wb.sheetnames:
        raise ValueError(f"Excel 缺少工作表: {name}")
    return wb[name]


def find_header_indexes(ws: Worksheet) -> Dict[str, int]:
    """
    扫描第1行，返回：列名 -> 列索引（1-based）
    """
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        headers[str(v).strip()] = col
    return headers


def ensure_columns(ws: Worksheet, headers: Dict[str, int], need_cols: List[str]) -> Dict[str, int]:
    """
    确保 need_cols 存在于表头，不存在则在末尾追加。返回更新后的列映射。
    """
    updated = dict(headers)
    for name in need_cols:
        if name not in updated:
            ws.cell(row=1, column=ws.max_column + 1, value=name)
            updated[name] = ws.max_column  # 刚写入的单元格已经生效
    return updated


def compact_preview(text: str, limit: int = 30) -> str:
    text = (text or "").replace("\n", " ").strip()
    return text if len(text) <= limit else text[:limit] + "..."


def extract_template_vars(template: str) -> List[str]:
    """
    从模板中提取所有 {{变量名}} 占位符，返回去重后的变量名列表。
    例如："查找{{学校名称}}在{{城市}}的信息" -> ["学校名称", "城市"]
    """
    pattern = r'\{\{([^}]+)\}\}'
    matches = re.findall(pattern, template)
    # 去除空格并去重，保持顺序
    seen = set()
    result = []
    for m in matches:
        name = m.strip()
        if name and name not in seen:
            seen.add(name)
            result.append(name)
    return result


def fill_template(template: str, values: Dict[str, Any]) -> str:
    """
    用字典值填充模板中的 {{变量名}} 占位符。
    例如：template="查找{{学校名称}}的信息", values={"学校名称": "北京一中"} 
         -> "查找北京一中的信息"
    """
    result = template
    for key, val in values.items():
        placeholder = f"{{{{{key}}}}}"
        result = result.replace(placeholder, str(val))
    return result


def is_empty_value(val: Any) -> bool:
    """
    判断是否为空值：None、空字符串、纯空格
    """
    if val is None:
        return True
    if isinstance(val, str) and val.strip() == "":
        return True
    return False


def is_json_array_text(s: str) -> bool:
    s = s.strip()
    return s.startswith("[") and s.endswith("]")


def extract_json_array_from_text(s: str) -> str:
    """
    兼容模型把 JSON 放在 ```json ... ``` 或前后有说明文字的情况。
    策略：
      1) 去除 ```...``` 包裹
      2) 从文本中找到最外层方括号的 JSON 段
    """
    text = s.strip()

    # 去除 ```json ... ``` 包裹
    fence = re.compile(r"^```(?:json|JSON)?\s*(.*?)\s*```$", re.S)
    m = fence.match(text)
    if m:
        text = m.group(1).strip()

    if is_json_array_text(text):
        return text

    # 宽松：从首个 '[' 到最后一个 ']' 的包裹
    lb = text.find("[")
    rb = text.rfind("]")
    if lb != -1 and rb != -1 and rb > lb:
        candidate = text[lb:rb + 1].strip()
        if is_json_array_text(candidate):
            return candidate

    # 失败则返回原文（让上层报错）
    return text


def call_llm_genai(
    client: genai.Client,
    model: str,
    system_prompt: str,
    user_content: str,
    timeout: int,
    tools: Optional[List[types.Tool]] = None,
    temperature: Optional[float] = None,
    thinking_budget: Optional[int] = None,
    debug: bool = False,
) -> Tuple[Optional[List[Dict[str, Any]]], Dict[str, Any], Optional[str]]:
    """
    调用 Google GenAI SDK 的 generate_content 接口。
    返回：(json_array或None, usage字典, 错误文本或None)
    
    参数:
        client: GenAI 客户端
        model: 模型 ID
        system_prompt: 系统提示词
        user_content: 用户内容
        timeout: 超时时间（秒）
        tools: 可选的工具列表（如 Google Search），用于启用联网搜索等功能
        temperature: 温度参数，控制随机性（0.0-2.0）
        thinking_budget: 思考预算，-1 表示无限制
        debug: 是否启用调试模式，打印请求和响应内容
    
    注意：timeout 参数保留在函数签名中以保持接口一致性，
    但 Google GenAI SDK 的 generate_content 可能不直接支持该参数。
    超时控制可能需要通过 Client 配置或其他方式实现。
    """
    try:
        # 构建配置对象
        # 新版 SDK 要求通过 GenerateContentConfig 传递所有配置参数
        config_kwargs = {}
        
        # 添加 system_instruction
        if system_prompt:
            config_kwargs["system_instruction"] = system_prompt
        
        # 添加 tools（如果提供）
        if tools:
            config_kwargs["tools"] = tools
        
        # 添加 temperature（如果提供）
        if temperature is not None:
            config_kwargs["temperature"] = temperature
        
        # 添加 thinking_config（如果提供）
        if thinking_budget is not None:
            config_kwargs["thinking_config"] = types.ThinkingConfig(
                thinking_budget=thinking_budget
            )
        
        # 创建配置对象（如果有任何配置）
        config = types.GenerateContentConfig(**config_kwargs) if config_kwargs else None
        
        # 调试模式：打印请求信息
        if debug:
            log("=" * 60)
            log("📤 API 请求详情")
            log("=" * 60)
            log(f"模型: {model}")
            log(f"系统提示 (前200字): {compact_preview(system_prompt, 200) if system_prompt else '(无)'}")
            log(f"用户内容 (前200字): {compact_preview(user_content, 200)}")
            if tools:
                log(f"工具: {[str(t) for t in tools]}")
            log("=" * 60)
        
        # 调用 API
        response = client.models.generate_content(
            model=model,
            contents=user_content,
            config=config
        )
        
    except Exception as e:
        if debug:
            log("=" * 60)
            log("❌ 请求异常")
            log("=" * 60)
            log(f"错误: {type(e).__name__}: {e}")
            import traceback
            log(f"堆栈:\n{traceback.format_exc()}")
            log("=" * 60)
        return None, {}, f"请求异常: {type(e).__name__}: {e}"

    # 提取响应文本
    try:
        content = response.text
    except Exception as e:
        return None, {}, f"响应缺少 text 属性: {type(e).__name__}: {e}"

    # 调试模式：打印响应信息
    if debug:
        log("=" * 60)
        log("📥 API 响应详情")
        log("=" * 60)
        log(f"原始响应 (前500字): {compact_preview(content, 500)}")
        
        # 检查 grounding metadata（联网搜索信息）
        if hasattr(response, 'candidates') and response.candidates:
            candidate = response.candidates[0]
            if hasattr(candidate, 'grounding_metadata') and candidate.grounding_metadata:
                metadata = candidate.grounding_metadata
                log(f"🌐 联网搜索信息:")
                if hasattr(metadata, 'web_search_queries') and metadata.web_search_queries:
                    log(f"  搜索查询: {metadata.web_search_queries}")
                if hasattr(metadata, 'grounding_chunks') and metadata.grounding_chunks:
                    log(f"  搜索结果数: {len(metadata.grounding_chunks)}")
                    for i, chunk in enumerate(metadata.grounding_chunks[:3], 1):
                        if hasattr(chunk, 'web') and chunk.web:
                            title = getattr(chunk.web, 'title', 'N/A')
                            uri = getattr(chunk.web, 'uri', 'N/A')
                            log(f"    {i}. {title}: {uri}")

    # 提取 usage 信息（如果存在）
    usage = {}
    try:
        # 尝试多种可能的 usage 属性路径
        if hasattr(response, 'usage_metadata') and response.usage_metadata:
            usage_meta = response.usage_metadata
            usage = {
                "prompt_tokens": getattr(usage_meta, 'prompt_token_count', 0) or 0,
                "completion_tokens": getattr(usage_meta, 'completion_token_count', 0) or 0,
                "total_tokens": getattr(usage_meta, 'total_token_count', 0) or 0,
            }
        elif hasattr(response, 'usage') and response.usage:
            # 兼容其他可能的 usage 格式
            usage_obj = response.usage
            usage = {
                "prompt_tokens": getattr(usage_obj, 'prompt_tokens', 0) or getattr(usage_obj, 'input_tokens', 0) or 0,
                "completion_tokens": getattr(usage_obj, 'completion_tokens', 0) or getattr(usage_obj, 'output_tokens', 0) or 0,
                "total_tokens": getattr(usage_obj, 'total_tokens', 0) or 0,
            }
        
        if debug and usage:
            log(f"📊 Token 使用: prompt={usage.get('prompt_tokens', 0)}, "
                f"completion={usage.get('completion_tokens', 0)}, "
                f"total={usage.get('total_tokens', 0)}")
    except Exception:
        # 如果无法提取 usage，继续执行（usage 为空字典）
        pass

    # 解析 JSON 数组
    content = extract_json_array_from_text(str(content))
    
    if debug:
        log(f"提取的 JSON (前500字): {compact_preview(content, 500)}")
    
    try:
        arr = json.loads(content)
    except Exception as e:
        if debug:
            log(f"❌ JSON 解析失败: {type(e).__name__}: {e}")
            log("=" * 60)
        return None, usage, f"内容不是 JSON 数组: {type(e).__name__}: {e}; 原文片段: {content[:1000]}"

    if not isinstance(arr, list):
        if debug:
            log(f"❌ 顶层不是数组，而是: {type(arr)}")
            log("=" * 60)
        return None, usage, "顶层非数组"
    
    # 元素必须为对象
    for i, it in enumerate(arr):
        if not isinstance(it, dict):
            if debug:
                log(f"❌ 数组第 {i+1} 个元素不是对象")
                log("=" * 60)
            return None, usage, f"数组第 {i+1} 个元素不是对象"
    
    if debug:
        log(f"✅ 成功解析 JSON 数组，包含 {len(arr)} 个元素")
        if arr:
            log(f"第一个元素的键: {list(arr[0].keys())}")
        log("=" * 60)
    
    return arr, usage, None


def with_retry(func, retry_times: int, retry_delay: int):
    def wrapper(*args, **kwargs):
        last_err = None
        for i in range(retry_times + 1):
            result = func(*args, **kwargs)
            # 约定：func 返回 (arr, usage, err_text)
            if result[2] is None:
                return result
            last_err = result[2]
            # 对可重试错误做简单判断（含 429/5xx 文本时退避），否则也简单等一等
            time.sleep(retry_delay if i < retry_times else 0)
        return (None, {}, last_err)
    return wrapper


def save_with_backup_atomic(wb, xlsx_path: str, made_backup: List[bool]) -> None:
    """
    首次保存前做 .bak 备份；使用临时文件 + 替换 的基本原子写法
    """
    if not made_backup[0]:
        bak = xlsx_path + ".bak"
        if not os.path.exists(bak):
            try:
                with open(xlsx_path, "rb") as rf, open(bak, "wb") as wf:
                    wf.write(rf.read())
                log(f"已创建备份: {bak}")
            except Exception as e:
                log(f"创建备份失败（忽略）: {e}")
        made_backup[0] = True

    tmp = xlsx_path + ".tmp"
    wb.save(tmp)
    # Windows 下替换
    try:
        if os.path.exists(xlsx_path):
            os.replace(tmp, xlsx_path)
        else:
            os.rename(tmp, xlsx_path)
    except Exception as e:
        log(f"保存替换失败: {e}")
        # 兜底直接写原文件（可能失败）
        wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="批量调用 LLM 并写回 Excel（Google GenAI SDK 版本）")
    parser.add_argument("--input-file", required=True, help="输入 Excel 文件路径")
    parser.add_argument("--config", default="config.toml", help="配置文件路径，默认 config.toml")
    parser.add_argument("--llm", required=True, help="使用的模型配置名，例如 genai_2_5_flash_latest")
    parser.add_argument("--rows", default=None, help="处理行范围，例如 2-5 或 2+；缺省处理全部")
    parser.add_argument("--api-key", default=None, help="可选；命令行覆盖配置中的 api_key")
    parser.add_argument("--debug", action="store_true", help="启用调试模式，输出详细日志")
    args = parser.parse_args()
    
    # 如果启用调试模式，配置日志
    if args.debug:
        import logging
        logging.basicConfig(
            level=logging.DEBUG,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        # 为相关的 logger 设置 DEBUG 级别
        for logger_name in ['google', 'google_genai', 'httpx', 'httpcore']:
            logging.getLogger(logger_name).setLevel(logging.DEBUG)
        log("已启用调试模式")

    xlsx_path = args.input_file
    if not os.path.exists(xlsx_path):
        print(f"找不到输入文件: {xlsx_path}", file=sys.stderr)
        sys.exit(2)

    # 读配置
    cfg = load_config(args.config)
    llm_cfg = merge_llm_config(cfg, args.llm, args.api_key)

    api_key = llm_cfg.get("api_key") or ""
    model_id = llm_cfg["model_id"]
    parallel = int(llm_cfg.get("parallel", 5))
    retry_times = int(llm_cfg.get("retry_times", 1))
    retry_delay = int(llm_cfg.get("retry_delay", 10))
    timeout = int(llm_cfg.get("timeout", 120))
    price_in = float(llm_cfg.get("price_per_1m_input_tokens", 0.0))
    price_out = float(llm_cfg.get("price_per_1m_output_tokens", 0.0))

    api_base = llm_cfg.get("api_base")
    # 空字符串视为未设置
    if api_base is not None and str(api_base).strip() == "":
        api_base = None
    
    enable_google_search = bool(llm_cfg.get("enable_google_search", False))
    enable_url_context = bool(llm_cfg.get("enable_url_context", True))
    
    # 生成参数（可选）
    temperature = llm_cfg.get("temperature")  # None 或浮点数
    if temperature is not None:
        temperature = float(temperature)
    
    thinking_budget = llm_cfg.get("thinking_budget")  # None 或整数
    if thinking_budget is not None:
        thinking_budget = int(thinking_budget)
    
    log("启动参数：")
    log(f"- input-file: {xlsx_path}")
    log(f"- llm: {args.llm}")
    log(f"- model_id: {model_id}")
    log(f"- api_key: {mask_key_tail(api_key)}")
    if api_base:
        log(f"- api_base: {api_base}")
    else:
        log(f"- api_base: (使用默认 Google API)")
    log(f"- parallel: {parallel}, retry_times: {retry_times}, retry_delay: {retry_delay}s, timeout: {timeout}s")
    log(f"- enable_google_search: {enable_google_search}")
    log(f"- enable_url_context: {enable_url_context}")
    if temperature is not None:
        log(f"- temperature: {temperature}")
    if thinking_budget is not None:
        log(f"- thinking_budget: {thinking_budget}")
    if args.rows:
        log(f"- rows: {args.rows}")

    # 创建 Google GenAI 客户端
    try:
        # 如果提供了 api_base，使用 http_options 自定义端点
        if api_base:
            client = genai.Client(
                api_key=api_key,
                http_options=types.HttpOptions(base_url=api_base)
            )
        else:
            client = genai.Client(api_key=api_key)
    except Exception as e:
        print(f"无法创建 GenAI 客户端：{e}", file=sys.stderr)
        sys.exit(2)

    # 创建工具（联网搜索和 URL Context）
    tools = None
    tools_list = []
    
    if enable_google_search:
        try:
            # 创建 Google Search 工具（使用 google_search 而不是 google_search_retrieval）
            # API 要求使用 google_search，而不是已弃用的 google_search_retrieval
            google_search = types.GoogleSearch()
            google_search_tool = types.Tool(google_search=google_search)
            tools_list.append(google_search_tool)
            log("✓ 已启用 Google 联网搜索功能")
        except Exception as e:
            log(f"⚠ 创建 Google Search 工具失败: {e}，将不使用联网搜索")
    
    if enable_url_context:
        try:
            # 创建 URL Context 工具，允许模型直接访问和理解网页内容
            # 注意：类名是 UrlContext（驼峰命名），不是 URLContext
            url_context = types.UrlContext()
            url_context_tool = types.Tool(url_context=url_context)
            tools_list.append(url_context_tool)
            log("✓ 已启用 URL Context 功能（模型可直接访问网页内容）")
        except Exception as e:
            log(f"⚠ 创建 URL Context 工具失败: {e}，将不使用 URL Context")
    
    if tools_list:
        tools = tools_list

    # 读 Excel
    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        print(f"无法打开 Excel：{e}", file=sys.stderr)
        sys.exit(2)

    ws_prompt = get_sheet(wb, "prompt")
    ws_qa = get_sheet(wb, "QA")

    # 读取 prompt 表（第1行表头，第2行内容）
    prompt_headers = find_header_indexes(ws_prompt)
    if "system" not in prompt_headers or "user" not in prompt_headers:
        print("prompt 表缺少必需列：system 或 user", file=sys.stderr)
        sys.exit(2)
    
    col_system = prompt_headers["system"]
    col_user = prompt_headers["user"]
    
    sys_prompt = ws_prompt.cell(row=2, column=col_system).value
    user_template = ws_prompt.cell(row=2, column=col_user).value
    
    if sys_prompt is None or str(sys_prompt).strip() == "":
        print("prompt 表的 system 列（第2行）不能为空", file=sys.stderr)
        sys.exit(2)
    if user_template is None or str(user_template).strip() == "":
        print("prompt 表的 user 列（第2行）不能为空", file=sys.stderr)
        sys.exit(2)
    
    sys_prompt = str(sys_prompt)
    user_template = str(user_template)
    
    # 从用户模板中提取输入字段
    input_fields = extract_template_vars(user_template)
    if not input_fields:
        print("user 模板中未找到任何 {{变量名}} 占位符", file=sys.stderr)
        sys.exit(2)
    
    log(f"从 user 模板中提取到 {len(input_fields)} 个输入字段: {input_fields}")

    # QA 表头
    headers = find_header_indexes(ws_qa)
    
    # 验证输入字段是否都在 QA 表中
    missing_fields = [f for f in input_fields if f not in headers]
    if missing_fields:
        print(f"QA 表缺少模板所需的输入字段: {missing_fields}", file=sys.stderr)
        sys.exit(2)
    
    # 确保控制列存在
    headers = ensure_columns(ws_qa, headers, [COL_FOUND, COL_ERROR])
    col_found = headers[COL_FOUND]
    col_err = headers[COL_ERROR]
    
    # 输入字段列映射
    input_cols = {field: headers[field] for field in input_fields}
    
    # 输出字段集合：表头中除去输入字段、FOUND、ERROR 的其它列（仅写这些）
    excluded = set(input_fields) | {COL_FOUND, COL_ERROR}
    output_cols = {k: v for k, v in headers.items() if k not in excluded}

    data_start_row = 2
    data_end_row_initial = ws_qa.max_row  # 启动时的原始末行（用于 rows 范围）
    target_rows = parse_rows_arg(args.rows, data_start_row, data_end_row_initial)

    # 统计去重：基于输入字段组合元组
    input_tuples_all = []
    input_tuples_done_set = set()
    for r in target_rows:
        # 读取所有输入字段值
        values = {}
        skip_row = False
        for field, col in input_cols.items():
            val = ws_qa.cell(row=r, column=col).value
            if is_empty_value(val):
                skip_row = True
                break
            values[field] = str(val)
        
        if skip_row:
            continue
        
        # 创建元组作为唯一标识
        tuple_key = tuple(values[field] for field in input_fields)
        input_tuples_all.append(tuple_key)
        
        # 检查是否已完成
        found_v = ws_qa.cell(row=r, column=col_found).value
        if not is_empty_value(found_v):
            input_tuples_done_set.add(tuple_key)
    
    unique_inputs = set(input_tuples_all)
    log(f"输入组合去重统计：总 {len(unique_inputs)} 组，其中已完成 {len(input_tuples_done_set)} 组")

    # 为 rows 范围执行插入偏移跟踪：记录"原始主行" -> 插入的额外行数
    inserted_below: Dict[int, int] = {}

    made_backup = [False]

    # 简单的进度累计
    total = len(target_rows)
    n_done = 0
    n_success = 0  # 有结果
    n_empty = 0    # 数组空
    n_error = 0

    # 费用累计（当 usage 存在时）
    sum_prompt_tokens = 0
    sum_completion_tokens = 0

    retry_call = with_retry(
        lambda *a, **kw: call_llm_genai(*a, **kw),
        retry_times=retry_times,
        retry_delay=retry_delay,
    )

    def current_row_pos(original_row: int) -> int:
        """根据已插入情况，计算该原始行现在的实际行号"""
        shift = 0
        for r0, added in inserted_below.items():
            if r0 < original_row:
                shift += added
        return original_row + shift

    for idx, r0 in enumerate(target_rows, start=1):
        r = current_row_pos(r0)
        
        # 读取所有输入字段值
        input_values = {}
        empty_fields = []
        for field, col in input_cols.items():
            val = ws_qa.cell(row=r, column=col).value
            if is_empty_value(val):
                empty_fields.append(field)
            else:
                input_values[field] = str(val)
        
        # 生成字段预览字符串（用于日志）
        preview_parts = [f"{field}='{compact_preview(input_values.get(field, ''), 20)}'" 
                        for field in input_fields]
        input_preview = "[" + ", ".join(preview_parts) + "]"
        
        # 判定是否跳过（断点续跑：FOUND 非空就跳过）
        found_val = ws_qa.cell(row=r, column=col_found).value
        if not is_empty_value(found_val):
            n_done += 1
            log(f"{idx}/{total} 跳过（已完成） r={r} {input_preview}")
            continue

        # 验证输入字段不能为空
        if empty_fields:
            error_msg = f"输入字段为空: {', '.join(empty_fields)}"
            ws_qa.cell(row=r, column=col_found, value="错误")
            ws_qa.cell(row=r, column=col_err, value=error_msg)
            save_with_backup_atomic(wb, xlsx_path, made_backup)
            n_done += 1
            n_error += 1
            log(f"{idx}/{total} 错误 r={r} {input_preview} -> {error_msg}")
            continue

        # 填充用户模板
        user_content = fill_template(user_template, input_values)
        
        # 请求
        arr, usage, err = retry_call(
            client, model_id, sys_prompt, user_content, timeout, tools,
            temperature, thinking_budget, args.debug
        )

        if usage:
            sum_prompt_tokens += int(usage.get("prompt_tokens", 0))
            sum_completion_tokens += int(usage.get("completion_tokens", 0))

        if err is not None:
            # 写入主行错误
            ws_qa.cell(row=r, column=col_found, value="错误")
            ws_qa.cell(row=r, column=col_err, value=str(err)[:500])
            save_with_backup_atomic(wb, xlsx_path, made_backup)
            n_done += 1
            n_error += 1
            log(f"{idx}/{total} 错误 r={r} {input_preview} -> {err}")
            continue

        # arr 一定是 list[dict]
        if len(arr) == 0:
            # 无结果：主行写"否"，不插入新行
            ws_qa.cell(row=r, column=col_found, value="否")
            ws_qa.cell(row=r, column=col_err, value="")
            # 清空输出列
            for name, c in output_cols.items():
                ws_qa.cell(row=r, column=c, value="")
            save_with_backup_atomic(wb, xlsx_path, made_backup)
            inserted_below[r0] = 0
            n_done += 1
            n_empty += 1
            log(f"{idx}/{total} 空结果 r={r} {input_preview}（已落盘）")
            continue

        # 有结果：主行写第1个，下面插入 len(arr)-1 行写其余
        # 关键：在修改主行之前，先读取原行的所有列值（用于拷贝到展开行）
        extra = max(0, len(arr) - 1)
        row_values = {}
        if extra > 0:
            # 先保存主行的所有原始列值
            for col_idx in range(1, ws_qa.max_column + 1):
                cell_value = ws_qa.cell(row=r, column=col_idx).value
                row_values[col_idx] = cell_value
        
        # 现在修改主行：只更新输出字段和控制字段，保持其他字段不变
        ws_qa.cell(row=r, column=col_found, value="是")
        ws_qa.cell(row=r, column=col_err, value="")
        # 写输出字段（只写入 JSON 中存在的字段，避免覆盖原有数据）
        first_obj = arr[0]
        for name, c in output_cols.items():
            if name in first_obj:  # 只有 JSON 中存在该字段时才写入
                v = first_obj[name]
                if isinstance(v, (dict, list)):
                    v = json.dumps(v, ensure_ascii=False)
                ws_qa.cell(row=r, column=c, value=v)

        # 插入展开行
        if extra > 0:
            ws_qa.insert_rows(r + 1, amount=extra)
            # 逐条写入
            for i in range(extra):
                rr = r + 1 + i
                # 先拷贝主行的所有列值
                for col_idx, value in row_values.items():
                    ws_qa.cell(row=rr, column=col_idx, value=value)
                
                # 然后覆盖控制字段
                ws_qa.cell(row=rr, column=col_found, value="是")
                ws_qa.cell(row=rr, column=col_err, value="")
                
                # 最后覆盖输出字段（写入新结果，只写入 JSON 中存在的字段）
                obj = arr[1 + i]
                for name, c in output_cols.items():
                    if name in obj:  # 只有 JSON 中存在该字段时才写入
                        v = obj[name]
                        if isinstance(v, (dict, list)):
                            v = json.dumps(v, ensure_ascii=False)
                        ws_qa.cell(row=rr, column=c, value=v)

        inserted_below[r0] = extra
        save_with_backup_atomic(wb, xlsx_path, made_backup)
        n_done += 1
        n_success += 1
        log(f"{idx}/{total} 成功 r={r} {input_preview} 展开 {len(arr)} 行（已落盘）")

    # 结束统计
    cost = 0.0
    if price_in or price_out:
        cost = (sum_prompt_tokens / 1_000_000.0) * price_in + (sum_completion_tokens / 1_000_000.0) * price_out

    log("完成。")
    log(f"- 总计：{total}, 成功(有结果)={n_success}, 空结果={n_empty}, 错误={n_error}")
    if (sum_prompt_tokens + sum_completion_tokens) > 0:
        log(f"- tokens: prompt={sum_prompt_tokens}, completion={sum_completion_tokens}, 估算费用=${cost:.4f}（按配置单价）")


if __name__ == "__main__":
    main()

