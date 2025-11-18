#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
LLM 批量调用：读取 Excel（prompt & QA），按模板逐行构造提示并请求 LLM，
将 JSON 数组结果展开写回原文件。满足：
- prompt 表：system 列存放系统提示，user 列存放用户提示模板（含 {{field}} 变量）
- QA 表：从各列读取数据填充模板变量，生成用户提示
- 展开结果的所有行：输入字段与 FOUND 相同
- 每个输入处理完成后立即落盘
- 支持 rows 范围、断点续跑、输入字段组合去重统计
"""

import argparse
import datetime as dt
import json
import os
import re
import sys
import time
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# --- 列名配置 ---
COL_FOUND = "FOUND"  # 结果状态列名
COL_ERROR = "ERROR"      # 错误信息列名

# --- 配置解析（tomllib 优先） ---
try:
    import tomllib  # py311+
except ModuleNotFoundError:
    try:
        import tomli as tomllib  # py310-
    except Exception:
        tomllib = None


def log(msg: str) -> None:
    now = dt.datetime.now().strftime("%H:%M:%S")
    print(f"[{now}] {msg}", flush=True)


def mask_key_tail(key: Optional[str]) -> str:
    if not key:
        return "(empty)"
    tail = key[-5:] if len(key) >= 5 else key
    return "*" * max(0, len(key) - 5) + tail


def load_config(path: str) -> dict:
    if not os.path.exists(path):
        raise FileNotFoundError(f"配置文件不存在: {path}")
    if tomllib is None:
        raise RuntimeError("缺少 tomllib/tomli，请安装 tomli 或使用 Python 3.11+")
    with open(path, "rb") as f:
        return tomllib.load(f)


def merge_llm_config(cfg: dict, llm_name: str, cli_api_key: Optional[str]) -> dict:
    base = cfg.get("llm", {}) or {}
    # TOML 中 [llm.gemini_search] 会被解析为嵌套表 cfg["llm"]["gemini_search"]
    # 支持两种写法：
    # 1) [llm] + [llm.gemini_search] - 标准嵌套表
    # 2) cfg.get("llm.gemini_search", {}) - 容错（某些解析器可能支持）
    llm_section = cfg.get("llm", {}) or {}
    if isinstance(llm_section, dict):
        by_table = llm_section.get(llm_name, {})  # 标准方式：cfg["llm"]["gemini_search"]
    else:
        by_table = {}
    # 容错：尝试直接键访问（某些 TOML 解析器可能支持）
    if not by_table:
        by_table = cfg.get(f"llm.{llm_name}", {}) or {}

    # 合并：base <- by_table
    merged = dict(base)
    if isinstance(by_table, dict):
        merged.update(by_table)

    # CLI api_key 优先
    if cli_api_key:
        merged["api_key"] = cli_api_key

    # 必要字段检查
    api_base = merged.get("api_base")
    api_key = merged.get("api_key")
    model_id = merged.get("model_id")
    if not api_base:
        raise ValueError("配置缺少 llm.api_base")
    if not (api_key or merged.get("user_token")):
        raise ValueError("未提供 api_key（或 user_token）")
    if not model_id:
        raise ValueError("配置缺少 llm.model_id")

    # 默认并发/重试/超时
    merged.setdefault("parallel", 5)
    merged.setdefault("retry_times", 1)
    merged.setdefault("retry_delay", 10)
    merged.setdefault("timeout", 120)
    return merged


def parse_rows_arg(rows_arg: Optional[str], data_start_row: int, data_end_row: int) -> List[int]:
    """
    rows 语法：
      - None: 处理 data_start_row..data_end_row
      - "2-5": 处理 2..5
      - "2+":  处理 2..data_end_row
    返回：原始行号列表（基于启动时的行号）
    """
    if not rows_arg:
        return list(range(data_start_row, data_end_row + 1))

    rows_arg = rows_arg.strip()
    m = re.fullmatch(r"(\d+)\-(\d+)", rows_arg)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        a = max(a, data_start_row)
        b = min(b, data_end_row)
        if a > b:
            return []
        return list(range(a, b + 1))

    m = re.fullmatch(r"(\d+)\+", rows_arg)
    if m:
        a = int(m.group(1))
        a = max(a, data_start_row)
        return list(range(a, data_end_row + 1))

    raise ValueError(f"rows 参数不合法: {rows_arg}")


def get_sheet(wb, name: str) -> Worksheet:
    if name not in wb.sheetnames:
        raise ValueError(f"Excel 缺少工作表: {name}")
    return wb[name]


def find_header_indexes(ws: Worksheet) -> Dict[str, int]:
    """
    扫描第1行，返回：列名 -> 列索引（1-based）
    """
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        headers[str(v).strip()] = col
    return headers


def ensure_columns(ws: Worksheet, headers: Dict[str, int], need_cols: List[str]) -> Dict[str, int]:
    """
    确保 need_cols 存在于表头，不存在则在末尾追加。返回更新后的列映射。
    """
    updated = dict(headers)
    for name in need_cols:
        if name not in updated:
            ws.cell(row=1, column=ws.max_column + 1, value=name)
            updated[name] = ws.max_column  # 刚写入的单元格已经生效
    return updated


def compact_preview(text: str, limit: int = 30) -> str:
    text = (text or "").replace("\n", " ").strip()
    return text if len(text) <= limit else text[:limit] + "..."


def extract_template_variables(template: str) -> List[str]:
    """
    从模板中提取所有 {{variable}} 格式的变量名。
    返回：去重后的变量名列表
    """
    pattern = r'\{\{([^}]+)\}\}'
    matches = re.findall(pattern, template)
    # 去除空格并去重，保持顺序
    seen = set()
    result = []
    for m in matches:
        name = m.strip()
        if name and name not in seen:
            seen.add(name)
            result.append(name)
    return result


def fill_template(template: str, values: Dict[str, str]) -> str:
    """
    用字典中的值填充模板中的 {{variable}} 占位符。
    """
    result = template
    for key, value in values.items():
        placeholder = f"{{{{{key}}}}}"
        result = result.replace(placeholder, str(value))
    return result


def is_json_array_text(s: str) -> bool:
    s = s.strip()
    return s.startswith("[") and s.endswith("]")


def extract_json_array_from_text(s: str) -> str:
    """
    兼容模型把 JSON 放在 ```json ... ``` 或前后有说明文字的情况。
    策略：
      1) 去除 ```...``` 包裹
      2) 从文本中找到最外层方括号的 JSON 段
    """
    text = s.strip()

    # 去除 ```json ... ``` 包裹
    fence = re.compile(r"^```(?:json|JSON)?\s*(.*?)\s*```$", re.S)
    m = fence.match(text)
    if m:
        text = m.group(1).strip()

    if is_json_array_text(text):
        return text

    # 宽松：从首个 '[' 到最后一个 ']' 的包裹
    lb = text.find("[")
    rb = text.rfind("]")
    if lb != -1 and rb != -1 and rb > lb:
        candidate = text[lb:rb + 1].strip()
        if is_json_array_text(candidate):
            return candidate

    # 失败则返回原文（让上层报错）
    return text


def call_llm_chat_openai_compatible(
    api_base: str,
    api_key: str,
    model: str,
    messages: List[Dict[str, str]],
    timeout: int,
) -> Tuple[Optional[List[Dict[str, Any]]], Dict[str, Any], Optional[str]]:
    """
    调用 OpenAI 兼容的 /chat/completions 接口。
    返回：(json_array或None, usage字典, 错误文本或None)
    """
    url = api_base.rstrip("/") + "/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "messages": messages,
        "stream": False,
        "temperature": 0.5,
    }
    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=timeout)
    except Exception as e:
        return None, {}, f"请求异常: {type(e).__name__}: {e}"

    if resp.status_code // 100 != 2:
        return None, {}, f"HTTP {resp.status_code}: {resp.text[:1000]}"

    try:
        data = resp.json()
    except Exception:
        return None, {}, f"响应非 JSON: {resp.text[:1000]}"

    usage = data.get("usage", {}) or {}

    try:
        content = data["choices"][0]["message"]["content"]
    except Exception:
        return None, usage, f"响应缺少 choices[0].message.content: {json.dumps(data)[:1000]}"

    content = extract_json_array_from_text(str(content))
    try:
        arr = json.loads(content)
    except Exception as e:
        return None, usage, f"内容不是 JSON 数组: {type(e).__name__}: {e}; 原文片段: {content[:1000]}"

    if not isinstance(arr, list):
        return None, usage, "顶层非数组"
    # 元素必须为对象
    for i, it in enumerate(arr):
        if not isinstance(it, dict):
            return None, usage, f"数组第 {i+1} 个元素不是对象"
    return arr, usage, None


def with_retry(func, retry_times: int, retry_delay: int):
    def wrapper(*args, **kwargs):
        last_err = None
        for i in range(retry_times + 1):
            result = func(*args, **kwargs)
            # 约定：func 返回 (arr, usage, err_text)
            if result[2] is None:
                return result
            last_err = result[2]
            # 对可重试错误做简单判断（含 429/5xx 文本时退避），否则也简单等一等
            time.sleep(retry_delay if i < retry_times else 0)
        return (None, {}, last_err)
    return wrapper


def save_with_backup_atomic(wb, xlsx_path: str, made_backup: List[bool]) -> None:
    """
    首次保存前做 .bak 备份；使用临时文件 + 替换 的基本原子写法
    """
    if not made_backup[0]:
        bak = xlsx_path + ".bak"
        if not os.path.exists(bak):
            try:
                with open(xlsx_path, "rb") as rf, open(bak, "wb") as wf:
                    wf.write(rf.read())
                log(f"已创建备份: {bak}")
            except Exception as e:
                log(f"创建备份失败（忽略）: {e}")
        made_backup[0] = True

    tmp = xlsx_path + ".tmp"
    wb.save(tmp)
    # Windows 下替换
    try:
        if os.path.exists(xlsx_path):
            os.replace(tmp, xlsx_path)
        else:
            os.rename(tmp, xlsx_path)
    except Exception as e:
        log(f"保存替换失败: {e}")
        # 兜底直接写原文件（可能失败）
        wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="批量调用 LLM 并写回 Excel")
    parser.add_argument("--input-file", required=True, help="输入 Excel 文件路径")
    parser.add_argument("--config", default="config.toml", help="配置文件路径，默认 config.toml")
    parser.add_argument("--llm", required=True, help="使用的模型配置名，例如 gemini_search")
    parser.add_argument("--rows", default=None, help="处理行范围，例如 2-5 或 2+；缺省处理全部")
    parser.add_argument("--api-key", default=None, help="可选；命令行覆盖配置中的 api_key")
    args = parser.parse_args()

    xlsx_path = args.input_file
    if not os.path.exists(xlsx_path):
        print(f"找不到输入文件: {xlsx_path}", file=sys.stderr)
        sys.exit(2)

    # 读配置
    cfg = load_config(args.config)
    llm_cfg = merge_llm_config(cfg, args.llm, args.api_key)

    api_base = llm_cfg["api_base"]
    api_key = llm_cfg.get("api_key") or ""
    model_id = llm_cfg["model_id"]
    parallel = int(llm_cfg.get("parallel", 5))
    retry_times = int(llm_cfg.get("retry_times", 1))
    retry_delay = int(llm_cfg.get("retry_delay", 10))
    timeout = int(llm_cfg.get("timeout", 120))
    price_in = float(llm_cfg.get("price_per_1m_input_tokens", 0.0))
    price_out = float(llm_cfg.get("price_per_1m_output_tokens", 0.0))

    log("启动参数：")
    log(f"- input-file: {xlsx_path}")
    log(f"- llm: {args.llm}")
    log(f"- api_base: {api_base}")
    log(f"- model_id: {model_id}")
    log(f"- api_key: {mask_key_tail(api_key)}")
    log(f"- parallel: {parallel}, retry_times: {retry_times}, retry_delay: {retry_delay}s, timeout: {timeout}s")
    if args.rows:
        log(f"- rows: {args.rows}")

    # 读 Excel
    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        print(f"无法打开 Excel：{e}", file=sys.stderr)
        sys.exit(2)

    ws_prompt = get_sheet(wb, "prompt")
    ws_qa = get_sheet(wb, "QA")

    # 读取 prompt 表的表头和数据（第1行=表头，第2行=数据）
    prompt_headers = find_header_indexes(ws_prompt)
    if "system" not in prompt_headers:
        print("prompt 表缺少 system 列", file=sys.stderr)
        sys.exit(2)
    if "user" not in prompt_headers:
        print("prompt 表缺少 user 列", file=sys.stderr)
        sys.exit(2)
    
    col_prompt_system = prompt_headers["system"]
    col_prompt_user = prompt_headers["user"]
    
    # 读取第2行数据
    sys_prompt = ws_prompt.cell(row=2, column=col_prompt_system).value
    user_template = ws_prompt.cell(row=2, column=col_prompt_user).value
    
    if sys_prompt is None or str(sys_prompt).strip() == "":
        print("prompt 表 system 列第2行不能为空", file=sys.stderr)
        sys.exit(2)
    if user_template is None or str(user_template).strip() == "":
        print("prompt 表 user 列第2行不能为空", file=sys.stderr)
        sys.exit(2)
    
    sys_prompt = str(sys_prompt).strip()
    user_template = str(user_template).strip()
    
    # 提取模板中的所有输入字段
    input_fields = extract_template_variables(user_template)
    if not input_fields:
        print("user 模板中没有找到任何 {{变量}}，至少需要一个输入字段", file=sys.stderr)
        sys.exit(2)
    
    log(f"用户提示模板: {compact_preview(user_template, 60)}")
    log(f"提取到的输入字段: {input_fields}")

    # QA 表头
    headers = find_header_indexes(ws_qa)
    
    # 验证所有输入字段都存在于 QA 表中
    missing_fields = [f for f in input_fields if f not in headers]
    if missing_fields:
        print(f"QA 表缺少模板所需的字段列: {missing_fields}", file=sys.stderr)
        sys.exit(2)
    
    # 确保必要列
    headers = ensure_columns(ws_qa, headers, [COL_FOUND, COL_ERROR])
    col_found = headers[COL_FOUND]
    col_err = headers[COL_ERROR]
    
    # 输入字段列索引
    input_cols = {field: headers[field] for field in input_fields}
    
    # 输出字段集合：表头中除去输入字段/FOUND/ERROR 的其它列（仅写这些）
    excluded = set(input_fields) | {COL_FOUND, COL_ERROR}
    output_cols = {k: v for k, v in headers.items() if k not in excluded}
    
    log(f"输入字段列: {list(input_cols.keys())}")
    log(f"输出字段列: {list(output_cols.keys())}")

    data_start_row = 2
    data_end_row_initial = ws_qa.max_row  # 启动时的原始末行（用于 rows 范围）
    target_rows = parse_rows_arg(args.rows, data_start_row, data_end_row_initial)

    # 统计去重：基于所有输入字段的组合值
    # 构造 key: 将所有输入字段值拼接为元组
    def make_input_key(row: int) -> Optional[Tuple[str, ...]]:
        values = []
        for field in input_fields:
            col_idx = input_cols[field]
            val = ws_qa.cell(row=row, column=col_idx).value
            val_str = "" if val is None else str(val).strip()
            values.append(val_str)
        # 如果所有字段都为空，则视为无效行
        if all(v == "" for v in values):
            return None
        return tuple(values)
    
    keys_all = []
    keys_done_set = set()
    for r in target_rows:
        key = make_input_key(r)
        if key is None:
            continue
        keys_all.append(key)
        found_v = ws_qa.cell(row=r, column=col_found).value
        if found_v is not None and str(found_v).strip() != "":
            keys_done_set.add(key)
    
    keys_all_unique = set(keys_all)
    log(f"候选输入组合去重统计：总 {len(keys_all_unique)} 个唯一输入，其中已完成 {len(keys_done_set)}")

    # 为 rows 范围执行插入偏移跟踪：记录“原始主行” -> 插入的额外行数
    inserted_below: Dict[int, int] = {}

    made_backup = [False]

    # 简单的进度累计
    total = len(target_rows)
    n_done = 0
    n_success = 0  # 有结果
    n_empty = 0    # 数组空
    n_error = 0

    # 费用累计（当 usage 存在时）
    sum_prompt_tokens = 0
    sum_completion_tokens = 0

    retry_call = with_retry(
        lambda *a, **kw: call_llm_chat_openai_compatible(*a, **kw),
        retry_times=retry_times,
        retry_delay=retry_delay,
    )

    def current_row_pos(original_row: int) -> int:
        """根据已插入情况，计算该原始行现在的实际行号"""
        shift = 0
        for r0, added in inserted_below.items():
            if r0 < original_row:
                shift += added
        return original_row + shift

    for idx, r0 in enumerate(target_rows, start=1):
        r = current_row_pos(r0)
        
        # 读取所有输入字段的值
        input_values = {}
        missing_or_empty = []
        for field in input_fields:
            col_idx = input_cols[field]
            val = ws_qa.cell(row=r, column=col_idx).value
            val_str = "" if val is None else str(val).strip()
            input_values[field] = val_str
            if val_str == "":
                missing_or_empty.append(field)
        
        # 构造输入字段预览（用于日志）
        input_preview = ", ".join([f"{k}='{compact_preview(v, 20)}'" for k, v in input_values.items()])
        
        # 判定是否跳过（断点续跑：主行 FOUND 非空就跳过）
        found_val = ws_qa.cell(row=r, column=col_found).value
        if found_val is not None and str(found_val).strip() != "":
            n_done += 1
            log(f"{idx}/{total} 跳过（已完成） 行{r} [{input_preview}]")
            continue

        # 验证输入字段不能为空
        if missing_or_empty:
            err_msg = f"输入字段为空: {missing_or_empty}"
            ws_qa.cell(row=r, column=col_found, value="错误")
            ws_qa.cell(row=r, column=col_err, value=err_msg)
            save_with_backup_atomic(wb, xlsx_path, made_backup)
            n_done += 1
            n_error += 1
            log(f"{idx}/{total} 错误：{err_msg}（行{r}），已落盘")
            continue

        # 用输入字段值填充用户提示模板
        user_prompt = fill_template(user_template, input_values)
        
        # 组消息
        messages = [
            {"role": "system", "content": sys_prompt},
            {"role": "user", "content": user_prompt},
        ]

        # 请求
        arr, usage, err = retry_call(
            llm_cfg["api_base"], api_key, model_id, messages, timeout
        )

        if usage:
            sum_prompt_tokens += int(usage.get("prompt_tokens", 0))
            sum_completion_tokens += int(usage.get("completion_tokens", 0))

        if err is not None:
            # 写入主行错误
            ws_qa.cell(row=r, column=col_found, value="错误")
            ws_qa.cell(row=r, column=col_err, value=str(err)[:500])
            save_with_backup_atomic(wb, xlsx_path, made_backup)
            n_done += 1
            n_error += 1
            log(f"{idx}/{total} 错误 行{r} [{input_preview}] -> {err}")
            continue

        # arr 一定是 list[dict]
        if len(arr) == 0:
            # 无结果：主行写"否"，不插入新行
            ws_qa.cell(row=r, column=col_found, value="否")
            ws_qa.cell(row=r, column=col_err, value="")
            # 清空输出列
            for name, c in output_cols.items():
                ws_qa.cell(row=r, column=c, value="")
            save_with_backup_atomic(wb, xlsx_path, made_backup)
            inserted_below[r0] = 0
            n_done += 1
            n_empty += 1
            log(f"{idx}/{total} 空结果 行{r} [{input_preview}]（已落盘）")
            continue

        # 有结果：主行写第1个，下面插入 len(arr)-1 行写其余
        # 所有展开行的输入字段与 FOUND 相同
        # 主行已有输入字段值，只需写控制列和输出列
        ws_qa.cell(row=r, column=col_found, value="是")
        ws_qa.cell(row=r, column=col_err, value="")
        # 写输出字段
        first_obj = arr[0]
        for name, c in output_cols.items():
            v = first_obj.get(name, "")
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            ws_qa.cell(row=r, column=c, value=v)

        extra = max(0, len(arr) - 1)
        if extra > 0:
            ws_qa.insert_rows(r + 1, amount=extra)
            # 逐条写入
            for i in range(extra):
                rr = r + 1 + i
                # 复制所有输入字段的值
                for field, col_idx in input_cols.items():
                    ws_qa.cell(row=rr, column=col_idx, value=input_values[field])
                # 写控制列
                ws_qa.cell(row=rr, column=col_found, value="是")
                ws_qa.cell(row=rr, column=col_err, value="")
                # 写输出字段
                obj = arr[1 + i]
                for name, c in output_cols.items():
                    v = obj.get(name, "")
                    if isinstance(v, (dict, list)):
                        v = json.dumps(v, ensure_ascii=False)
                    ws_qa.cell(row=rr, column=c, value=v)

        inserted_below[r0] = extra
        save_with_backup_atomic(wb, xlsx_path, made_backup)
        n_done += 1
        n_success += 1
        log(f"{idx}/{total} 成功 行{r} [{input_preview}] 展开 {len(arr)} 行（已落盘）")

    # 结束统计
    cost = 0.0
    if price_in or price_out:
        cost = (sum_prompt_tokens / 1_000_000.0) * price_in + (sum_completion_tokens / 1_000_000.0) * price_out

    log("完成。")
    log(f"- 总计：{total}, 成功(有结果)={n_success}, 空结果={n_empty}, 错误={n_error}")
    if (sum_prompt_tokens + sum_completion_tokens) > 0:
        log(f"- tokens: prompt={sum_prompt_tokens}, completion={sum_completion_tokens}, 估算费用=${cost:.4f}（按配置单价）")


if __name__ == "__main__":
    main()
