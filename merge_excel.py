#!/usr/bin/env python3
import argparse
import glob
import os
import re
import shutil
from openpyxl import load_workbook


def parse_args():
    parser = argparse.ArgumentParser(
        description="合并通过 split_excel.py 拆分的 Excel 文件"
    )
    parser.add_argument(
        "--input-pattern",
        required=True,
        help="输入文件模式，例如：data/file-split-200.*.xlsx",
    )
    parser.add_argument(
        "--output-file",
        required=True,
        help="输出文件路径（.xlsx）",
    )
    parser.add_argument(
        "--sheet",
        required=True,
        help="要合并的 sheet 名称（区分大小写）",
    )
    parser.add_argument(
        "--header",
        type=int,
        default=1,
        help="表头行数，默认为 1（需与拆分时一致）",
    )
    return parser.parse_args()


def extract_part_number(filename: str) -> int:
    """
    从文件名中提取分片序号
    例如：data-split-200.1.xlsx -> 1
          data-split-200.2.xlsx -> 2
    """
    # 匹配类似 .数字.xlsx 的模式
    match = re.search(r"\.(\d+)\.xlsx$", filename, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return -1


def get_sorted_files(pattern: str) -> list:
    """
    根据模式匹配文件，并按分片序号排序
    """
    files = glob.glob(pattern)
    if not files:
        return []
    
    # 提取序号并排序
    file_with_numbers = []
    for f in files:
        num = extract_part_number(f)
        if num >= 0:
            file_with_numbers.append((num, f))
    
    # 按序号排序
    file_with_numbers.sort(key=lambda x: x[0])
    
    return [f for _, f in file_with_numbers]


def check_file_sequence(files: list) -> None:
    """
    检查文件序号是否连续
    """
    numbers = [extract_part_number(f) for f in files]
    
    if not numbers:
        return
    
    expected = list(range(1, len(numbers) + 1))
    if numbers != expected:
        print(f"[警告] 文件序号不连续！")
        print(f"  期望序号：{expected}")
        print(f"  实际序号：{numbers}")
        print(f"  将继续合并，但请检查是否缺失文件...")


def verify_headers_match(files: list, sheet_name: str, header_rows: int) -> bool:
    """
    严格验证所有文件的表头是否完全一致
    返回 True 表示所有表头一致，False 表示不一致
    """
    if not files:
        return True
    
    # 读取第一个文件的表头作为基准
    first_wb = load_workbook(files[0])
    if sheet_name not in first_wb.sheetnames:
        print(f"[错误] 第一个文件中未找到 sheet：{sheet_name}")
        first_wb.close()
        return False
    
    first_ws = first_wb[sheet_name]
    
    # 提取第一个文件的表头内容
    base_headers = []
    for row_idx in range(1, header_rows + 1):
        row_data = []
        for col_idx in range(1, first_ws.max_column + 1):
            cell = first_ws.cell(row_idx, col_idx)
            row_data.append(cell.value)
        base_headers.append(row_data)
    
    base_col_count = first_ws.max_column
    first_wb.close()
    
    print(f"[信息] 基准表头：{header_rows} 行 × {base_col_count} 列")
    
    # 验证其他文件的表头
    for i, file_path in enumerate(files[1:], start=2):
        wb = load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            print(f"[错误] 文件 {i} 中未找到 sheet：{sheet_name}")
            print(f"       文件路径：{file_path}")
            wb.close()
            return False
        
        ws = wb[sheet_name]
        
        # 检查列数
        if ws.max_column != base_col_count:
            print(f"[错误] 文件 {i} 的列数不匹配！")
            print(f"       基准列数：{base_col_count}")
            print(f"       当前列数：{ws.max_column}")
            print(f"       文件路径：{file_path}")
            wb.close()
            return False
        
        # 逐行逐列比较表头内容
        for row_idx in range(1, header_rows + 1):
            for col_idx in range(1, base_col_count + 1):
                cell = ws.cell(row_idx, col_idx)
                base_value = base_headers[row_idx - 1][col_idx - 1]
                
                if cell.value != base_value:
                    print(f"[错误] 文件 {i} 的表头内容不匹配！")
                    print(f"       位置：第 {row_idx} 行，第 {col_idx} 列")
                    print(f"       基准值：{base_value}")
                    print(f"       当前值：{cell.value}")
                    print(f"       文件路径：{file_path}")
                    wb.close()
                    return False
        
        wb.close()
    
    print(f"[验证] 所有 {len(files)} 个文件的表头完全一致 ✓")
    return True


def main():
    args = parse_args()
    
    input_pattern = args.input_pattern
    output_file = args.output_file
    sheet_name = args.sheet
    header_rows = args.header
    
    # 参数检查
    if header_rows < 0:
        print("[错误] 参数 --header 不能小于 0")
        return
    
    # 检查输出文件是否已存在
    if os.path.exists(output_file):
        print(f"[错误] 输出文件已存在：{output_file}")
        print("       请指定不同的 --output-file 或删除现有文件")
        return
    
    # 匹配并排序文件
    print(f"[信息] 搜索文件模式：{input_pattern}")
    files = get_sorted_files(input_pattern)
    
    if not files:
        print(f"[错误] 未找到匹配的文件")
        return
    
    print(f"[信息] 找到 {len(files)} 个文件：")
    for i, f in enumerate(files, 1):
        print(f"  {i}. {f}")
    
    # 检查文件序号连续性
    check_file_sequence(files)
    
    # 验证所有文件的 sheet 存在
    print(f"\n[验证] 检查所有文件是否包含 sheet '{sheet_name}'...")
    for f in files:
        try:
            wb = load_workbook(f)
            if sheet_name not in wb.sheetnames:
                print(f"[错误] 文件中未找到 sheet '{sheet_name}'：{f}")
                wb.close()
                return
            wb.close()
        except Exception as e:
            print(f"[错误] 无法打开文件：{f}")
            print(f"       错误信息：{e}")
            return
    
    print(f"[验证] 所有文件都包含 sheet '{sheet_name}' ✓")
    
    # 严格验证表头一致性
    print(f"\n[验证] 检查所有文件的表头是否完全一致...")
    if not verify_headers_match(files, sheet_name, header_rows):
        print("\n[错误] 表头验证失败，合并终止")
        return
    
    # 开始合并
    print(f"\n[合并] 开始合并文件...")
    
    # 复制第一个文件作为输出文件基础
    print(f"[步骤 1] 复制第一个文件作为基础：{files[0]}")
    shutil.copy2(files[0], output_file)
    
    # 打开输出文件
    output_wb = load_workbook(output_file)
    output_ws = output_wb[sheet_name]
    
    # 记录第一个文件的数据行数
    data_start_row = header_rows + 1
    current_max_row = output_ws.max_row
    first_file_data_rows = max(0, current_max_row - header_rows)
    
    print(f"         第一个文件数据行数：{first_file_data_rows}")
    
    total_data_rows = first_file_data_rows
    
    # 追加后续文件的数据行
    for i, file_path in enumerate(files[1:], start=2):
        print(f"[步骤 {i}] 追加文件：{file_path}")
        
        # 读取源文件
        source_wb = load_workbook(file_path)
        source_ws = source_wb[sheet_name]
        source_max_row = source_ws.max_row
        
        # 计算数据行范围
        if source_max_row <= header_rows:
            print(f"         警告：此文件没有数据行，跳过")
            source_wb.close()
            continue
        
        data_rows_count = source_max_row - header_rows
        
        # 逐行复制数据（跳过表头）
        for row_idx in range(data_start_row, source_max_row + 1):
            # 获取目标行号
            target_row = current_max_row + 1
            
            # 复制整行的单元格
            for col_idx in range(1, source_ws.max_column + 1):
                source_cell = source_ws.cell(row_idx, col_idx)
                target_cell = output_ws.cell(target_row, col_idx)
                
                # 复制值
                target_cell.value = source_cell.value
                
                # 可选：复制样式（如果需要）
                # target_cell.font = copy(source_cell.font)
                # target_cell.border = copy(source_cell.border)
                # target_cell.fill = copy(source_cell.fill)
                # target_cell.number_format = source_cell.number_format
                # target_cell.alignment = copy(source_cell.alignment)
            
            current_max_row += 1
        
        total_data_rows += data_rows_count
        print(f"         追加数据行数：{data_rows_count}")
        
        source_wb.close()
    
    # 保存输出文件
    output_wb.save(output_file)
    output_wb.close()
    
    print(f"\n[完成] 合并成功！")
    print(f"       输出文件：{output_file}")
    print(f"       总数据行数：{total_data_rows}")
    print(f"       总行数（含表头）：{header_rows + total_data_rows}")


if __name__ == "__main__":
    main()

