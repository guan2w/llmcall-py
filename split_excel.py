#!/usr/bin/env python3
import argparse
import math
import os
import shutil
from openpyxl import load_workbook


def parse_args():
    parser = argparse.ArgumentParser(
        description="按指定 sheet 和数据行数拆分 Excel 文件"
    )
    parser.add_argument(
        "--input-file",
        required=True,
        help="输入 Excel 文件路径（.xlsx）",
    )
    parser.add_argument(
        "--sheet",
        required=True,
        help="需要拆分的 sheet 名称（区分大小写）",
    )
    parser.add_argument(
        "--header",
        type=int,
        default=1,
        help="表头行数，0 表示没有表头，默认为 1",
    )
    parser.add_argument(
        "--each",
        type=int,
        required=True,
        help="每个拆分文件中包含的数据行数（不含表头）",
    )
    return parser.parse_args()


def build_output_filename(input_path: str, each: int, part_index: int) -> str:
    """
    将拆分信息插入到 .xlsx 之前：
    input:  /path/data.xlsx
    each:   100
    index:  1
    =>      /path/data-split-100.1.xlsx
    """
    dir_name, base_name = os.path.split(input_path)
    root, ext = os.path.splitext(base_name)

    # 只针对 .xlsx 做规范命名，其它后缀也尽量按类似规则处理
    if ext.lower() == ".xlsx":
        new_base = f"{root}-split-{each}.{part_index}{ext}"
    else:
        # 不标准情况，尽量保持原扩展名
        new_base = f"{root}-split-{each}.{part_index}{ext}"

    return os.path.join(dir_name, new_base)


def main():
    args = parse_args()

    input_file = args.input_file
    sheet_name = args.sheet
    header_rows = args.header
    each = args.each

    # 基本参数检查
    if not os.path.exists(input_file):
        print(f"[错误] 输入文件不存在：{input_file}")
        return

    if each <= 0:
        print("[错误] 参数 --each 必须为大于 0 的整数")
        return

    if header_rows < 0:
        print("[错误] 参数 --header 不能小于 0")
        return

    # 先打开一次原始文件，用来检查 sheet 和行数
    try:
        wb = load_workbook(input_file)
    except Exception as e:
        print(f"[错误] 打开 Excel 文件失败：{e}")
        return

    if sheet_name not in wb.sheetnames:
        print(f"[错误] 未找到指定 sheet：{sheet_name}")
        print("当前文件中的 sheet 名称：")
        for name in wb.sheetnames:
            print(f"  - {name}")
        return

    ws = wb[sheet_name]
    max_row = ws.max_row or 0

    # 计算数据起始行
    if header_rows == 0:
        data_start_row = 1
    else:
        data_start_row = header_rows + 1

    if data_start_row > max_row:
        print(
            f"[提示] 表头行数为 {header_rows}，总行数为 {max_row}，"
            "没有数据行可供拆分，脚本退出。"
        )
        return

    data_count = max_row - data_start_row + 1
    if data_count <= 0:
        print("[提示] 该 sheet 没有数据行可供拆分，脚本退出。")
        return

    # 计算需要拆分成多少份
    num_parts = math.ceil(data_count / each)

    print(f"[信息] 输入文件：{input_file}")
    print(f"[信息] 目标 sheet：{sheet_name}")
    print(f"[信息] 表头行数：{header_rows}")
    print(f"[信息] 数据行范围：{data_start_row} ~ {max_row}（共 {data_count} 行）")
    print(f"[信息] 每份数据行数：{each}")
    print(f"[信息] 将拆分为 {num_parts} 份")

    # 逐份拆分
    for i in range(num_parts):
        part_index = i + 1

        # 本份需要保留的数据行区间
        start_row = data_start_row + i * each
        end_row = min(start_row + each - 1, data_start_row + data_count - 1)

        out_file = build_output_filename(input_file, each, part_index)

        # 复制原始文件，然后在副本上删行
        shutil.copy2(input_file, out_file)

        try:
            part_wb = load_workbook(out_file)
        except Exception as e:
            print(f"[错误] 打开输出文件失败（{out_file}）：{e}")
            continue

        if sheet_name not in part_wb.sheetnames:
            print(f"[错误] 输出文件中未找到 sheet：{sheet_name}（{out_file}）")
            part_wb.close()
            continue

        part_ws = part_wb[sheet_name]
        part_max_row = part_ws.max_row or 0

        # 从底向上删除不在当前分片范围内的“数据行”
        for row in range(part_max_row, 0, -1):
            # header_rows == 0 时，data_start_row == 1，没有表头
            if row < data_start_row:
                # 表头行（如果有），不删除
                continue
            # 数据行中，如果不在当前分片范围内，则删除
            if row < start_row or row > end_row:
                part_ws.delete_rows(row, 1)

        # 保存修改
        part_wb.save(out_file)
        part_wb.close()

        print(
            f"[OK] 生成文件：{out_file} "
            f"(保留数据行 {start_row} ~ {end_row})"
        )

    wb.close()
    print("[完成] 拆分结束。")


if __name__ == "__main__":
    main()

